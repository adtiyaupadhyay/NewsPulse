"""
excel_export.py
------------------
Exports all stored articles into a formatted Excel file — with a
pivot-style summary sheet (category x sentiment counts) plus a
full data sheet. This is the "MS Excel for data analysis" piece
of the JD, done properly rather than just a raw CSV dump.

Why two sheets instead of one?
  - "Summary" sheet: what a manager/recruiter skims in 5 seconds —
    aggregated counts, ready for a quick read.
  - "All Articles" sheet: the full raw data, for anyone who wants
    to filter/sort/dig in themselves.
That split mirrors how real reporting tools present data — a top-level
view plus a drill-down, not just one giant table.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from database import get_all_articles


def export_to_excel(filename: str = "newspulse_report.xlsx"):
    articles = get_all_articles()
    if not articles:
        print("No articles to export — run the pipeline first.")
        return

    df = pd.DataFrame(articles)

    # Build a pivot-style summary: count of articles per category x sentiment
    summary = pd.pivot_table(
        df,
        index="category",
        columns="sentiment",
        values="id",
        aggfunc="count",
        fill_value=0,
    )

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary")
        df.to_excel(writer, sheet_name="All Articles", index=False)

        # --- Formatting pass: make it look like a real report, not a data dump ---
        workbook = writer.book

        for sheet_name in ["Summary", "All Articles"]:
            worksheet = workbook[sheet_name]

            # Bold header row with a light fill color
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            # Auto-width columns based on content length, so nothing
            # gets cut off or looks squished — a small thing recruiters
            # notice when opening a "professional-looking" file.
            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                    default=10,
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)

    print(f"Excel report saved: {filename}")
    print(f"  Summary sheet: category x sentiment breakdown")
    print(f"  All Articles sheet: {len(df)} rows")


if __name__ == "__main__":
    export_to_excel()